import openpyxl
import os

excel_file = "files/IS-004-01資產及風險評鑑表.xlsx"

def analyze_risk_sheet():
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb["資產及風險評鑑表"]
    
    print("--- 尋找標題列 ---")
    headers = []
    header_row_idx = 3 # based on previous observation
    for col in range(1, 20):
        val1 = sheet.cell(row=3, column=col).value
        val2 = sheet.cell(row=4, column=col).value
        
        # Combine merged headers conceptually
        header = f"{val1 or ''} {val2 or ''}".strip()
        headers.append(header)
    
    print(headers)
    
    print("--- 讀取風險評估資料行 ---")
    # Asset usually starts at row 5
    for i in range(5, 8):
        asset = sheet.cell(row=i, column=2).value # 名稱
        c = sheet.cell(row=i, column=6).value     # 機密性(C) likely around col 6/7/8?
        i_val = sheet.cell(row=i, column=7).value # 完整性(I)
        a = sheet.cell(row=i, column=8).value     # 可用性(A)
        # We need to find exactly which columns represent C, I, A
        print(f"Row {i} ASSET '{asset}' -> Columns 6-10: {[sheet.cell(row=i, column=c_idx).value for c_idx in range(6, 11)]}")

if __name__ == "__main__":
    analyze_risk_sheet()
