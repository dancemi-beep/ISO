import openpyxl
import os

def process_risk_assessment(input_file, output_file, acceptable_risk):
    # Load workbook and make sure formulas can be calculated if needed,
    # though openpyxl writes formulas as strings and Excel computes them on open.
    wb = openpyxl.load_workbook(input_file)
    sheet = wb["資產及風險評鑑表"]
    
    print(f"Processing Excel: {input_file} (可接受風險值設定為: {acceptable_risk})")
    
    # 模拟填寫三個資產的威脅(T), 弱點(V), 衝擊(I) 等級
    # Col 12=P, Col 21=T, Col 23=V, Col 24=I, Col 25=Risk
    
    # Asset 1 (Row 5 - 資訊人員) (P=5)
    sheet.cell(row=5, column=21).value = 2  # T
    sheet.cell(row=5, column=23).value = 2  # V
    sheet.cell(row=5, column=24).value = 3  # I
    sheet.cell(row=5, column=25).value = "=L5*U5*W5*X5" # Risk = 5 * 2 * 2 * 3 = 60
    
    # Asset 2 (Row 6 - 開發人員) (P=4)
    sheet.cell(row=6, column=21).value = 1  # T
    sheet.cell(row=6, column=23).value = 1  # V
    sheet.cell(row=6, column=24).value = 1  # I
    sheet.cell(row=6, column=25).value = "=L6*U6*W6*X6" # Risk = 4 * 1 * 1 * 1 = 4
    
    # Check if Risk > threshold
    # Openpyxl won't evaluate the formula we just wrote. In a real system we'd 
    # either calculate it in python or rely on win32com/xlwings if we must read it back.
    # Here we do the math in python just for the analytics report generation.
    
    # Use data_only=True to read cached values from formulas
    wb_data = openpyxl.load_workbook(input_file, data_only=True)
    sheet_data = wb_data["資產及風險評鑑表"]
    
    assets = []
    # Read the data back to classify
    for r in range(5, 7):
        name = sheet.cell(row=r, column=2).value
        p = sheet_data.cell(row=r, column=12).value or 0
        t = sheet.cell(row=r, column=21).value or 0
        v = sheet.cell(row=r, column=23).value or 0
        i_val = sheet.cell(row=r, column=24).value or 0
        
        try:
            # handle cases where cell.value is string or None or empty
            p_val = float(p) if str(p).strip() else 0.0
            t_val = float(t) if str(t).strip() else 0.0
            v_val = float(v) if str(v).strip() else 0.0
            i_val_float = float(i_val) if str(i_val).strip() else 0.0
            
            risk_score = p_val * t_val * v_val * i_val_float
            if risk_score > 0:
                assets.append({
                    "name": name,
                    "score": risk_score
                })
        except Exception as e:
            print(f"Skipping {name}: {e}")
            
    wb.save(output_file)
    print(f"Saved processed Excel to {output_file}")
    
    print("\n--- 風險評估結果 (Risk Assessment Analytics) ---")
    high_risks = [a for a in assets if a['score'] > acceptable_risk]
    
    print(f"總資產數: {len(assets)}")
    print(f"高於可接受風險值 ({acceptable_risk}) 的高風險資產數: {len(high_risks)}")
    for hr in high_risks:
        print(f" - [警告] {hr['name']} (風險值: {hr['score']}) => 需要填寫風險處理計畫！")

if __name__ == "__main__":
    template = "files/IS-004-01資產及風險評鑑表.xlsx"
    output = "output/IS-004-01_calculated.xlsx"
    
    if not os.path.exists("output"):
        os.makedirs("output")
        
    # 模擬使用者在問卷階段輸入「可接受風險值」為 25
    user_input_acceptable_risk = 25
    process_risk_assessment(template, output, acceptable_risk=user_input_acceptable_risk)
