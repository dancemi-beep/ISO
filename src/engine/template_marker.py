"""
Inserts a visible '⚠ 待修改' notice at the top of operational forms
so users know which sections require their manual review.
"""
import os
from docx import Document
from docx.shared import Pt, RGBColor, Cm, Emu
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml.ns import qn
from docx.oxml import OxmlElement
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# Files that need the "user action required" notice
OPERATIONAL_DOCX = [
    "qpower-002-02內部稽核查檢表.docx",
    "qpower-002-03矯正與預防處理單.docx",
    "qpower-002-08內部稽核報告.docx",
    "qpower-002-資訊安全管理程序.docx",
    "qpower-006-01帳號審查表.docx",
    "qpower-006-02資訊服務申請單.docx",
    "qpower-007-01資訊機房進出管制表.docx",
    "qpower-008-01備份管理作業表.docx",
    "qpower-011-01 供應商評鑑表.docx",
    "qpower-012-02資訊安全事故報告單.docx",
    "qpower-013-01業務流程營運衝擊分析表.docx",
    "qpower-013-02 營運持續計畫演練規劃暨處理報告單.docx",
    "qpower-016-01變更申請表.docx",
    "qpower-017-01雲端服務安全控制措施查檢表.docx",
    # Also tag the ones with residual XXX as reviewer name / record format
    "qpower-002-07內部稽核計畫.docx",
    "qpower-012-01威脅情資與監控活動管制表(範例).docx",
]

OPERATIONAL_XLSX = [
    "qpower-004-01資產及風險評鑑表.xlsx",
]

NOTICE_TEXT = (
    "⚠ 使用者須知：本文件為範本。"
    "標記為 □、___、XXX 的欄位需由貴公司依實際情況填寫或勾選。"
    "請務必逐一確認後再進行簽核發行。"
)


def add_notice_to_docx(filepath):
    """Insert a highlighted notice paragraph at the very top of a .docx file."""
    doc = Document(filepath)

    # Create notice paragraph and insert at position 0
    notice_para = OxmlElement('w:p')
    
    # Create the run
    run_elem = OxmlElement('w:r')
    
    # Run properties: bold, yellow highlight, larger font
    rPr = OxmlElement('w:rPr')
    
    bold = OxmlElement('w:b')
    rPr.append(bold)
    
    sz = OxmlElement('w:sz')
    sz.set(qn('w:val'), '24')  # 12pt
    rPr.append(sz)
    
    color = OxmlElement('w:color')
    color.set(qn('w:val'), 'FF6600')  # Orange
    rPr.append(color)
    
    highlight = OxmlElement('w:highlight')
    highlight.set(qn('w:val'), 'yellow')
    rPr.append(highlight)
    
    run_elem.append(rPr)
    
    # Text
    t = OxmlElement('w:t')
    t.set(qn('xml:space'), 'preserve')
    t.text = NOTICE_TEXT
    run_elem.append(t)
    
    notice_para.append(run_elem)
    
    # Insert at the very beginning of the document body
    doc.element.body.insert(0, notice_para)
    
    doc.save(filepath)
    print(f"  ✅ Marked: {os.path.basename(filepath)}")


def add_notice_to_xlsx(filepath):
    """Insert a highlighted notice row at the top of each sheet in an .xlsx file."""
    wb = openpyxl.load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.insert_rows(1)
        cell = ws.cell(row=1, column=1, value=NOTICE_TEXT)
        cell.font = Font(bold=True, color="FF6600", size=11)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True)
        # Merge across a reasonable number of columns
        max_col = min(ws.max_column or 1, 20)
        if max_col > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    
    wb.save(filepath)
    print(f"  ✅ Marked: {os.path.basename(filepath)}")


def mark_all_templates(marked_dir):
    """Apply notices to all operational form templates."""
    print("Marking operational forms with user notice...")
    
    for fname in OPERATIONAL_DOCX:
        fpath = os.path.join(marked_dir, fname)
        if os.path.exists(fpath):
            try:
                add_notice_to_docx(fpath)
            except Exception as e:
                print(f"  ❌ Error marking {fname}: {e}")
        else:
            print(f"  ⚠ Not found: {fname}")
    
    for fname in OPERATIONAL_XLSX:
        fpath = os.path.join(marked_dir, fname)
        if os.path.exists(fpath):
            try:
                add_notice_to_xlsx(fpath)
            except Exception as e:
                print(f"  ❌ Error marking {fname}: {e}")
        else:
            print(f"  ⚠ Not found: {fname}")
    
    print("Done.")


if __name__ == "__main__":
    mark_all_templates("files/marked")
